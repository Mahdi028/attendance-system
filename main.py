import cv2
import face_recognition
import numpy as np
from customtkinter import *
import pandas as pd
from datetime import datetime
import jdatetime
import os
import threading
from database import *
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from persiantools.jdatetime import JalaliDate
import time

current_mode = None
stop_camera = False
status_label = None


def boss_panel():
    try:
        if clock_job:
            app.after_cancel(clock_job)
    except Exception:
        pass
    
    for widget in app.winfo_children():
        widget.destroy()

    title = CTkLabel(app, text="ورود به پنل 🔐", font=("Arial", 30, "bold"))
    title.pack(pady=20)
    
    username_label = CTkLabel(app, text="نام کاربری را وارد کنید", font=("Arial", 16))
    username_label.pack(pady=10)
    
    username_entry = CTkEntry(app, width=200, font=("Arial", 16))
    username_entry.pack(pady=5)

    password_label = CTkLabel(app, text="رمز عبور را وارد کنید", font=("Arial", 16))
    password_label.pack(pady=10)

    password_entry = CTkEntry(app, show="*", width=200, font=("Arial", 16))
    password_entry.pack(pady=5)

    result_label = CTkLabel(app, text="", font=("Arial", 16))
    result_label.pack(pady=10)

    def check_password():
        username = username_entry.get().strip()
        password = password_entry.get().strip()

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute("SELECT username, password_hash FROM admin WHERE id=1")
        admin_data = cursor.fetchone()
        conn.close()

        if not admin_data:
            result_label.configure(text="❌ هیچ ادمینی یافت نشد.")
            return

        db_username, db_password_hash = admin_data

        if username != db_username:
            result_label.configure(text="❌ نام کاربری اشتباه است.")
        elif hash_password(password) != db_password_hash:
            result_label.configure(text="❌ رمز عبور اشتباه است.")
        else:
            result_label.configure(text="✅ ورود موفقیت‌آمیز")
            show_admin_panel()


    login_button = CTkButton(app, text="ورود", command=check_password, width=120, height=35)
    login_button.pack(pady=10)

    back_button = CTkButton(app, text="منو به بازگشت", command=show_main_menu, width=120, height=35)
    back_button.pack(pady=10)

def add_employees():
    try:
        if clock_job:
            app.after_cancel(clock_job)
    except Exception:
        pass
    
    for widget in app.winfo_children():
        widget.destroy()
    
    label1 = CTkLabel(app, text="افزودن کارمند", font=("Arial", 26, "bold"))
    label1.pack(pady=20)
    
    select_image_path = StringVar(value="")
    
    image_status_label = CTkLabel(app, text="هنوز هیچ تصویری انتخاب نشده", font=("Arial", 18))
    image_status_label.place(x=50, y=115)
    
    status_label = CTkLabel(app, text="", font=("Arial", 18))
    status_label.place(x=60, y=460)
    
    label2 = CTkLabel(app, text=": نام و نام خانوادگی کارمند", font=("Arial", 20))
    label2.place(x=400, y=200)
    employee_name_entry = CTkEntry(app, width=210, font=("Arial", 16))
    employee_name_entry.place(x=50, y=200)
    
    label3 = CTkLabel(app, text=": کد کارمند", font=("Arial", 20))
    label3.place(x=400, y=300)
    employee_id_entry = CTkEntry(app, width=210, font=("Arial", 16))
    employee_id_entry.place(x=50, y=300)
    
    def select_employee_image():
        file_path = filedialog.askopenfilename(
            title="انتخاب تصویر کارمند", 
            filetypes=[("تصاویر", "*.jpg;*.jpeg;*.png")]
        )
        if file_path:
            select_image_path.set(file_path)
            image_status_label.configure(text="✅ تصویر انتخاب شد", text_color="green")
        else:
            image_status_label.configure(text="❌ تصویر انتخاب نشد", text_color="red")

    def add_employees_successful():
        full_name = employee_name_entry.get().strip()
        code = employee_id_entry.get().strip()
        image_path = select_image_path.get()

        if not full_name or not code or not image_path:
            status_label.configure(text="⚠️ لطفاً همه فیلدها را پر کنید و تصویر انتخاب شود", text_color="red")
            return

        # --- ثبت در دیتابیس ---
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT INTO employees (full_name, code, image_path, date_added)
                VALUES (?, ?, ?, ?)
            """, (full_name, code, image_path, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit()
        except sqlite3.IntegrityError:
            status_label.configure(text="❌ این کد کارمند قبلاً ثبت شده است.", text_color="red")
            conn.close()
            return
        conn.close()

        # # --- مسیر پوشه اکسل ---
        # excel_dir = os.path.join(os.path.dirname(__file__), "Excel Employees")
        # os.makedirs(excel_dir, exist_ok=True)
        # excel_file_path = os.path.join(excel_dir, f"{full_name}.xlsx")

        # # --- ایجاد فایل اکسل ---
        # wb = Workbook()
        # ws = wb.active
        # ws.title = "حضور و غیاب"

        # # ستون‌ها
        # ws.append(["روز", "تاریخ", "ساعت ورود", "ساعت خروج"])

        # # تاریخ و روز امروز (شمسی)
        # today_jalali = JalaliDate.today()
        # weekday_map = {
        #     0: "شنبه",
        #     1: "یکشنبه",
        #     2: "دوشنبه",
        #     3: "سه‌شنبه",
        #     4: "چهارشنبه",
        #     5: "پنج‌شنبه",
        #     6: "جمعه"
        # }

        # day_name = weekday_map[today_jalali.weekday()]
        # date_str = today_jalali.strftime("%Y/%m/%d")

        # # افزودن یک ردیف نمونه (میتونی حذفش کنی)
        # ws.append([day_name, date_str, "", ""])

        # # تنظیم عرض ستون‌ها بر اساس طول داده
        # for column_cells in ws.columns:
        #     max_length = 0
        #     column_letter = get_column_letter(column_cells[0].column)
        #     for cell in column_cells:
        #         try:
        #             max_length = max(max_length, len(str(cell.value)))
        #         except:
        #             pass
        #     ws.column_dimensions[column_letter].width = max_length + 3

        # wb.save(excel_file_path)

        status_label.configure(text=f"✅ کارمند {full_name} با موفقیت اضافه شد", text_color="green")
        select_image_path.set("")
        image_status_label.configure(text="هنوز هیچ تصویری انتخاب نشده", text_color="black")
        employee_name_entry.delete(0, END)
        employee_id_entry.delete(0, END)

    select_image_btn = CTkButton(app, text="انتخاب تصویر کارمند", width=150, height=60, command=select_employee_image, font=("Arial", 14))
    select_image_btn.place(x=400, y=100)
    
    add_button = CTkButton(app, text="افزودن", width=120, height=35, command=add_employees_successful)
    add_button.place(x=400, y=380)
    
    back_button = CTkButton(app, text="منو به بازگشت", command=show_main_menu, width=120, height=35)
    back_button.place(x=90, y=380)


def delete_employee():
    try:
        if clock_job:
            app.after_cancel(clock_job)
    except Exception:
        pass
    
    for widget in app.winfo_children():
        widget.destroy()

    label1 = CTkLabel(app, text="حذف کارمند", font=("Arial", 26, "bold"))
    label1.pack(pady=20)

    label2 = CTkLabel(app, text=": نام کارمند را انتخاب کنید", font=("Arial", 20))
    label2.place(x=400, y=100)

    # دریافت نام‌ها از دیتابیس
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT full_name FROM employees")
    employees = [row[0] for row in cursor.fetchall()]
    conn.close()

    # اگر هیچ کارمندی وجود ندارد
    if not employees:
        CTkLabel(app, text="⚠️ هیچ کارمندی در دیتابیس وجود ندارد.", font=("Arial", 18)).place(x=200, y=200)
        back_button = CTkButton(app, text="بازگشت به منو", command=show_main_menu, width=120, height=35)
        back_button.place(x=260, y=300)
        return

    selected_employee = StringVar(value="اسامی کارمندان")

    employee_menu = CTkOptionMenu(app, values=employees, variable=selected_employee, width=200, font=("Arial", 14))
    employee_menu.place(x=60, y=100)

    status_label = CTkLabel(app, text="", font=("Arial", 18))
    status_label.place(x=60, y=250)

    def on_select(choice):
        status_label.configure(text=f"✅ کارمند «{choice}» برای حذف انتخاب شد", text_color="green")

    employee_menu.configure(command=on_select)

    def delete_selected_employee():
        choice = selected_employee.get()
        if choice == "اسامی کارمندان" or choice not in employees:
            status_label.configure(text="⚠️ لطفاً یک کارمند معتبر را انتخاب کنید", text_color="red")
            return

        # حذف از دیتابیس
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM employees WHERE full_name=?", (choice,))
        conn.commit()
        conn.close()

        # حذف از لیست و آپدیت منو
        employees.remove(choice)
        if employees:
            employee_menu.configure(values=employees)
            selected_employee.set("اسامی کارمندان")
            status_label.configure(text=f"🗑 کارمند «{choice}» حذف شد", text_color="red")
        else:
            status_label.configure(text=f"🗑 کارمند «{choice}» حذف شد. دیگر کارمندی باقی نمانده است", text_color="red")
            employee_menu.configure(values=["هیچ کارمندی نیست"])
            selected_employee.set("هیچ کارمندی نیست")

    delete_button = CTkButton(app, text="حذف", width=150, height=40, command=delete_selected_employee, font=("Arial", 14))
    delete_button.place(x=400, y=380)

    back_button = CTkButton(app, text="منو به بازگشت", command=show_main_menu, width=120, height=35)
    back_button.place(x=90, y=380)


def change_username():
    try:
        if clock_job:
            app.after_cancel(clock_job)
    except Exception:
        pass
    
    for widget in app.winfo_children():
        widget.destroy()
    
    label1 = CTkLabel(app, text="تغییر نام کاربری", font=("Arial", 26, "bold"))
    label1.pack(pady=20)
    
    label2 = CTkLabel(app, text=": نام کاربری فعلی", font=("Arial", 16))
    label2.pack(pady=10)
    
    last_username_entry = CTkEntry(app, width=200, font=("Arial", 16))
    last_username_entry.pack(pady=5)
    
    label3 = CTkLabel(app, text=": نام کاربری جدید", font=("Arial", 16))
    label3.pack(pady=10)
    
    new_username_entry = CTkEntry(app, width=200, font=("Arial", 16))
    new_username_entry.pack(pady=5)
    
    def change_username_successful():
        old_username = last_username_entry.get().strip()
        new_username = new_username_entry.get().strip()

        if not old_username or not new_username:
            CTkLabel(app, text="⚠️ لطفاً همه فیلدها را پر کنید.", font=("Arial", 16)).pack(pady=10)
            return
        if change_username_in_db(old_username, new_username):
            CTkLabel(app, text=f".تغییر یافت {new_username} نام کاربری جدید با موفقیت به", font=("Arial", 16)).pack(pady=10)
            check_btn.configure(state="disabled", text="✅ شد انجام تغییر")
        else:
            CTkLabel(app, text="❌ نام کاربری فعلی اشتباه است یا خطایی رخ داد.", font=("Arial", 16)).pack(pady=10)

    check_btn = CTkButton(app, text="تغییر", command=change_username_successful, width=120, height=35)
    check_btn.pack(pady=10)
    
    back_button = CTkButton(app, text="منو به بازگشت", command=show_main_menu, width=120, height=35)
    back_button.pack(pady=10)

def change_password():
    try:
        if clock_job:
            app.after_cancel(clock_job)
    except Exception:
        pass
    
    for widget in app.winfo_children():
        widget.destroy()
    
    label1 = CTkLabel(app, text="تغییر رمز عبور", font=("Arial", 26, "bold"))
    label1.pack(pady=20)
    
    label2 = CTkLabel(app, text=": رمز عبور فعلی", font=("Arial", 16))
    label2.pack(pady=10)
    
    last_password_entry = CTkEntry(app, width=200, font=("Arial", 16))
    last_password_entry.pack(pady=5)
    
    label3 = CTkLabel(app, text=": رمز عبور جدید", font=("Arial", 16))
    label3.pack(pady=10)
    
    new_password_entry = CTkEntry(app, width=200, font=("Arial", 16))
    new_password_entry.pack(pady=5)
    
    def change_password_successful():
        old_password = last_password_entry.get()
        new_password = new_password_entry.get()

        if not old_password or not new_password:
            CTkLabel(app, text="⚠️ لطفاً همه فیلدها را پر کنید.", font=("Arial", 16)).pack(pady=10)
            return

        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()

        cursor.execute("SELECT password_hash FROM admin WHERE id=1")
        current = cursor.fetchone()

        if not current:
            CTkLabel(app, text="❌ خطا در خواندن اطلاعات ادمین.", font=("Arial", 16)).pack(pady=10)
            conn.close()
            return

        stored_hash = current[0]

        if hash_password(old_password) != stored_hash:
            CTkLabel(app, text="❌ رمز عبور فعلی اشتباه است.", font=("Arial", 16)).pack(pady=10)
        elif hash_password(new_password) == stored_hash:
            CTkLabel(app, text="⚠️ رمز عبور جدید با رمز قبلی یکسان است.", font=("Arial", 16)).pack(pady=10)
        else:
            cursor.execute("UPDATE admin SET password_hash=?, last_updated=? WHERE id=1",
                        (hash_password(new_password), datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit()
            CTkLabel(app, text=f"رمز عبور جدید با موفقیت به {new_password} تغییر یافت.", font=("Arial", 16)).pack(pady=10)
            check_btn.configure(state="disabled", text="✅ شد انجام تغییر")

        conn.close()

    check_btn = CTkButton(app, text="تغییر", command=change_password_successful, width=120, height=35)
    check_btn.pack(pady=10)
    
    back_button = CTkButton(app, text="منو به بازگشت", command=show_main_menu, width=120, height=35)
    back_button.pack(pady=10)


def show_admin_panel():
    try:
        if clock_job:
            app.after_cancel(clock_job)
    except Exception:
        pass
    
    for widget in app.winfo_children():
        widget.destroy()

    label = CTkLabel(app, text="پنل مدیر 👨‍💼", font=("Arial", 26, "bold"))
    label.pack(pady=20)

    add_employee_btn = CTkButton(app, text="➕ افزودن کارمند", width=200, height=40, command=add_employees)
    add_employee_btn.pack(pady=10)
    
    add_employee_btn = CTkButton(app, text="❌ کارمند حذف", width=200, height=40, command=delete_employee)
    add_employee_btn.pack(pady=10)

    download_excel_btn = CTkButton(app, text="📥 اکسل فایل‌های دانلود", width=200, height=40)
    download_excel_btn.pack(pady=10)
    
    change_username_btn = CTkButton(app, text="🔑 کاربری نام تغییر", width=200, height=40, command=change_username)
    change_username_btn.pack(pady=10)

    change_passsord_btn = CTkButton(app, text="🔑 عبور رمز تغییر", width=200, height=40, command=change_password)
    change_passsord_btn.pack(pady=10)

    back_button = CTkButton(app, text="⬅ منو به بازگشت", command=show_main_menu, width=200, height=40)
    back_button.pack(pady=20)

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from persiantools.jdatetime import JalaliDate
from datetime import datetime
import os

def update_excel(full_name, mode):
    """
    ثبت ورود یا خروج در فایل اکسل کارمند با:
    - تاریخ شمسی
    - روز هفته فارسی
    - تنظیم خودکار عرض ستون‌ها
    - فونت و ترازبندی مناسب
    """
    excel_dir = os.path.join(os.path.dirname(__file__), "Excel Employees")
    os.makedirs(excel_dir, exist_ok=True)
    excel_path = os.path.join(excel_dir, f"{full_name}.xlsx")

    now = datetime.now()
    time_str = now.strftime("%H:%M:%S")

    today_j = JalaliDate.today()
    date_str = f"{today_j.day}/{today_j.month}/{today_j.year}"

    weekday_map = {
        0: "شنبه",
        1: "یکشنبه",
        2: "دوشنبه",
        3: "سه‌شنبه",
        4: "چهارشنبه",
        5: "پنج‌شنبه",
        6: "جمعه"
    }
    day_name = weekday_map[today_j.weekday()]

    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "حضور و غیاب"
        ws.append(["روز", "تاریخ", "ساعت ورود", "ساعت خروج"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        wb.save(excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active

    today_row = None
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[1]) == date_str:
            today_row = idx
            break

    if today_row:
        entry_cell = ws[f"C{today_row}"]
        exit_cell = ws[f"D{today_row}"]
    else:
        ws.append([day_name, date_str, "", ""])
        today_row = ws.max_row
        entry_cell = ws[f"C{today_row}"]
        exit_cell = ws[f"D{today_row}"]

    if mode == "entry" and not entry_cell.value:
        entry_cell.value = time_str
    elif mode == "exit" and entry_cell.value and not exit_cell.value:
        exit_cell.value = time_str
    else:
        wb.save(excel_path)
        return False

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 3

    wb.save(excel_path)
    return True


# def update_excel(full_name, mode):
#     """ثبت ورود یا خروج در فایل اکسل کارمند"""
#     excel_dir = os.path.join(os.path.dirname(__file__), "Excel Employees")
#     os.makedirs(excel_dir, exist_ok=True)
#     excel_path = os.path.join(excel_dir, f"{full_name}.xlsx")

#     # --- تاریخ و زمان فعلی ---
#     now = datetime.now()
#     today_jalali = JalaliDate.today()
#     date_str = today_jalali.strftime("%Y/%m/%d")
#     time_str = now.strftime("%H:%M:%S")

#     # --- روز هفته ---
#     weekday_map = {
#         0: "شنبه",
#         1: "یکشنبه",
#         2: "دوشنبه",
#         3: "سه‌شنبه",
#         4: "چهارشنبه",
#         5: "پنج‌شنبه",
#         6: "جمعه"
#     }
#     day_name = weekday_map[today_jalali.weekday()]

#     # # --- اگر فایل وجود نداشت، بساز ---
#     if not os.path.exists(excel_path):
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "حضور و غیاب"
#         ws.append(["روز", "تاریخ", "ساعت ورود", "ساعت خروج"])
#         wb.save(excel_path)

#     # --- باز کردن فایل موجود ---
#     wb = load_workbook(excel_path)
#     ws = wb.active

#     # --- جستجوی تاریخ امروز در فایل ---
#     today_row = None
#     for row in ws.iter_rows(min_row=2, values_only=True):
#         if str(row[1]) == date_str:
#             today_row = row
#             break

#     if today_row:
#         row_index = [r[1] for r in ws.iter_rows(min_row=2)].index(ws[f"B{ws.max_row}"].value) + 2
#     else:
#         # --- اضافه کردن ردیف جدید برای امروز ---
#         ws.append([day_name, date_str, "", ""])
#         row_index = ws.max_row

#     # --- خواندن مقادیر فعلی ---
#     entry_cell = ws[f"C{row_index}"]
#     exit_cell = ws[f"D{row_index}"]

#     # --- بروزرسانی ورود یا خروج ---
#     if mode == "entry" and not entry_cell.value:
#         entry_cell.value = time_str
#     elif mode == "exit" and entry_cell.value and not exit_cell.value:
#         exit_cell.value = time_str
#     else:
#         wb.save(excel_path)
#         return False  # تکراری یا غیرمجاز

#     # --- تنظیم عرض ستون‌ها ---
#     for column_cells in ws.columns:
#             max_length = 0
#             column_letter = get_column_letter(column_cells[0].column)
#             for cell in column_cells:
#                 try:
#                     max_length = max(max_length, len(str(cell.value)))
#                 except:
#                     pass
#             ws.column_dimensions[column_letter].width = max_length + 3

#     wb.save(excel_path)
#     return True

def enter_cv2():
    # # حذف لیبل‌های اضافی
    # for widget in app.winfo_children():
    #     if isinstance(widget, CTkLabel):
    #         widget.destroy()
    
    global current_mode, stop_camera, status_label
    if current_mode == "entry":
        stop_camera = True
        cv2.destroyAllWindows()
        time.sleep(1)
    current_mode = "exit"
    stop_camera = False


    if status_label is None or not status_label.winfo_exists():
        status_label = CTkLabel(app, text="در حال بارگذاری چهره‌ها برای ثبت ورود...", font=("Arial", 18))
        status_label.pack(pady=20)
    else:
        status_label.configure(text="در حال بارگذاری چهره‌ها...")
    app.update()

    # --- بارگذاری چهره‌ها از دیتابیس ---
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, full_name, image_path FROM employees")
    employees = cursor.fetchall()
    conn.close()

    if not employees:
        status_label.configure(text="❌ هیچ کارمندی در دیتابیس وجود ندارد.", text_color="red")
        return

    known_faces = []
    known_names = []
    known_ids = []

    for emp_id, name, img_path in employees:
        if img_path and os.path.exists(img_path):
            try:
                image = face_recognition.load_image_file(img_path)
                encoding = face_recognition.face_encodings(image)
                if encoding:
                    known_faces.append(encoding[0])
                    known_names.append(name)
                    known_ids.append(emp_id)
            except Exception as e:
                print(f"⚠️ خطا در بارگذاری تصویر {name}: {e}")

    if not known_faces:
        status_label.configure(text="❌ هیچ چهره‌ای برای شناسایی وجود ندارد.", text_color="red")
        return

    status_label.configure(text="📷 دوربین فعال شد، برای ثبت ورود مقابل دوربین بایستید...", text_color="yellow")
    app.update()

    # --- فعال‌سازی دوربین ---
    video_capture = cv2.VideoCapture("http://192.168.1.4:8080/video")
    # video_capture = cv2.VideoCapture(0)
    process_frame = True
    detected_names = set()

    while True:
        if stop_camera:
            break

        ret, frame = video_capture.read()
        if not ret:
            status_label.configure(text="⚠️ خطا در دریافت تصویر از دوربین.", text_color="red")
            break

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        if process_frame:
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

            face_names = []

            for face_encoding in face_encodings:
                matches = face_recognition.compare_faces(known_faces, face_encoding)
                name = "Unknown"
                emp_id = None

                if True in matches:
                    match_index = np.argmin(face_recognition.face_distance(known_faces, face_encoding))
                    emp_id = known_ids[match_index]
                    name = known_names[match_index]

                    # جلوگیری از چند ثبت متوالی
                    if name in detected_names:
                        face_names.append(name)
                        continue

                    conn = sqlite3.connect(DB_PATH)
                    cursor = conn.cursor()
                    date_today = datetime.now().strftime("%Y-%m-%d")
                    time_now = datetime.now().strftime("%H:%M:%S")

                    cursor.execute("""
                        SELECT entry_time, exit_time FROM attendance
                        WHERE employee_id=? AND date=?
                    """, (emp_id, date_today))
                    record = cursor.fetchone()

                    if not record:
                        # ثبت ورود جدید
                        cursor.execute("""
                            INSERT INTO attendance (employee_id, date, entry_time)
                            VALUES (?, ?, ?)
                        """, (emp_id, date_today, time_now))
                        conn.commit()

                        # ثبت در اکسل کارمند
                        excel_dir = os.path.join(os.path.dirname(__file__), "Excel Employees")
                        os.makedirs(excel_dir, exist_ok=True)
                        excel_path = os.path.join(excel_dir, f"{name}.xlsx")
                        if os.path.exists(excel_path):
                            df = pd.read_excel(excel_path)
                        else:
                            df = pd.DataFrame(columns=["روز", "تاریخ", "ساعت ورود", "ساعت خروج"])

                        new_row = {
                            "روز": datetime.now().strftime("%A"),
                            "تاریخ": date_today,
                            "ساعت ورود": time_now,
                            "ساعت خروج": ""
                        }
                        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                        df.to_excel(excel_path, index=False)

                        status_label.configure(
                            text=f"✅ ورود {name} در ساعت {time_now} ثبت شد.",
                            text_color="green"
                        )
                        detected_names.add(name)
                    else:
                        entry_time, exit_time = record
                        if entry_time is not None:
                            status_label.configure(
                                text=f"⚠️ ورود {name} قبلاً در ساعت {entry_time} ثبت شده است.",
                                text_color="orange"
                            )
                        else:
                            cursor.execute("""
                                UPDATE attendance SET entry_time=? WHERE employee_id=? AND date=?
                            """, (time_now, emp_id, date_today))
                            conn.commit()
                            status_label.configure(
                                text=f"✅ ورود {name} در ساعت {time_now} ثبت شد.",
                                text_color="green"
                            )
                            detected_names.add(name)

                    conn.close()

                face_names.append(name)

        process_frame = not process_frame

        # --- رسم کادر و نام روی فریم ---
        for (top, right, bottom, left), name in zip(face_locations, face_names):
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
            cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), color, cv2.FILLED)
            cv2.putText(frame, name, (left + 6, bottom - 10),
                        cv2.FONT_HERSHEY_DUPLEX, 0.9, (255, 255, 255), 2)

        cv2.namedWindow("Enter Recognition", cv2.WINDOW_NORMAL)
        cv2.resizeWindow("Enter Recognition", 600, 550)
        cv2.moveWindow("Enter Recognition", 700, 100)
        cv2.imshow("Enter Recognition", frame)

        
        # cv2.resizeWindow("Enter Recognition", 600, 550)
        # cv2.imshow("Enter Recognition", frame)
        # cv2.moveWindow("Enter Recognition", 700, 100)
        app.update_idletasks()
        app.update()

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    video_capture.release()
    cv2.destroyAllWindows()
    status_label.configure(text="🛑 سیستم تشخیص چهره برای ورود متوقف شد.", text_color="red")

def exit_cv2():
    # # حذف لیبل‌های اضافی
    # for widget in app.winfo_children():
    #     if isinstance(widget, CTkLabel):
    #         widget.destroy()
    
    global current_mode, stop_camera, status_label
    if current_mode == "exit":
        stop_camera = True
        cv2.destroyAllWindows()
        time.sleep(1)
    current_mode = "entry"
    stop_camera = False

    if status_label is None or not status_label.winfo_exists():
        status_label = CTkLabel(app, text="در حال بارگذاری چهره‌ها برای ثبت خروج...", font=("Arial", 18))
        status_label.pack(pady=20)
    else:
        status_label.configure(text="در حال بارگذاری چهره‌ها...")
    app.update()

    # --- بارگذاری چهره‌ها از دیتابیس ---
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("SELECT id, full_name, image_path FROM employees")
    employees = cursor.fetchall()
    conn.close()

    if not employees:
        status_label.configure(text="❌ هیچ کارمندی در دیتابیس وجود ندارد.", text_color="red")
        return

    known_faces = []
    known_names = []
    known_ids = []

    for emp_id, name, img_path in employees:
        if img_path and os.path.exists(img_path):
            try:
                image = face_recognition.load_image_file(img_path)
                encoding = face_recognition.face_encodings(image)
                if encoding:
                    known_faces.append(encoding[0])
                    known_names.append(name)
                    known_ids.append(emp_id)
            except Exception as e:
                print(f"⚠️ خطا در بارگذاری تصویر {name}: {e}")

    if not known_faces:
        status_label.configure(text="❌ هیچ چهره‌ای برای شناسایی وجود ندارد.", text_color="red")
        return

    status_label.configure(text="📷 دوربین فعال شد، برای ثبت خروج مقابل دوربین بایستید...", text_color="yellow")
    app.update()

    # --- فعال‌سازی دوربین ---
    video_capture = cv2.VideoCapture("http://192.168.1.4:8080/video")
    # video_capture = cv2.VideoCapture(0)
    process_frame = True
    detected_names = set()

    while True:
        if stop_camera:
            break

        ret, frame = video_capture.read()
        if not ret:
            status_label.configure(text="⚠️ خطا در دریافت تصویر از دوربین.", text_color="red")
            break

        small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
        rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

        if process_frame:
            face_locations = face_recognition.face_locations(rgb_small_frame)
            face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

            face_names = []

            for face_encoding in face_encodings:
                matches = face_recognition.compare_faces(known_faces, face_encoding)
                name = "Unknown"
                emp_id = None

                if True in matches:
                    match_index = np.argmin(face_recognition.face_distance(known_faces, face_encoding))
                    emp_id = known_ids[match_index]
                    name = known_names[match_index]

                    # جلوگیری از چند ثبت متوالی
                    if name in detected_names:
                        face_names.append(name)
                        continue

                    conn = sqlite3.connect(DB_PATH)
                    cursor = conn.cursor()
                    date_today = datetime.now().strftime("%Y-%m-%d")
                    time_now = datetime.now().strftime("%H:%M:%S")

                    cursor.execute("""
                        SELECT entry_time, exit_time FROM attendance
                        WHERE employee_id=? AND date=?
                    """, (emp_id, date_today))
                    record = cursor.fetchone()

                    if not record:
                        status_label.configure(
                            text=f"⚠️ {name} امروز ورود ثبت نشده است، خروج قابل ثبت نیست.",
                            text_color="red"
                        )
                    else:
                        entry_time, exit_time = record
                        if exit_time is None:
                            # ثبت خروج
                            cursor.execute("""
                                UPDATE attendance
                                SET exit_time=?
                                WHERE employee_id=? AND date=?
                            """, (time_now, emp_id, date_today))
                            conn.commit()

                            # به‌روزرسانی اکسل شخص
                            excel_dir = os.path.join(os.path.dirname(__file__), "Excel Employees")
                            os.makedirs(excel_dir, exist_ok=True)
                            excel_path = os.path.join(excel_dir, f"{name}.xlsx")
                            if os.path.exists(excel_path):
                                df = pd.read_excel(excel_path)
                                mask = df["تاریخ"] == date_today
                                df.loc[mask, "ساعت خروج"] = time_now
                                df.to_excel(excel_path, index=False)

                            status_label.configure(
                                text=f"👋 خروج شما، {name} عزیز در ساعت {time_now} ثبت شد.",
                                text_color="blue"
                            )
                            detected_names.add(name)
                        else:
                            status_label.configure(
                                text=f"ℹ️ {name} امروز خروج قبلاً ثبت شده است.",
                                text_color="gray"
                            )

                    conn.close()

                face_names.append(name)

        process_frame = not process_frame

        # --- رسم کادر و نام روی فریم ---
        for (top, right, bottom, left), name in zip(face_locations, face_names):
            top *= 4
            right *= 4
            bottom *= 4
            left *= 4

            color = (0, 255, 0) if name != "Unknown" else (0, 0, 255)
            cv2.rectangle(frame, (left, top), (right, bottom), color, 2)
            cv2.rectangle(frame, (left, bottom - 35), (right, bottom), color, cv2.FILLED)
            cv2.putText(frame, name, (left + 6, bottom - 10),
                        cv2.FONT_HERSHEY_DUPLEX, 0.9, (255, 255, 255), 2)

        cv2.namedWindow("Exit Recognition", cv2.WINDOW_NORMAL)
        cv2.resizeWindow("Exit Recognition", 600, 550)
        cv2.moveWindow("Exit Recognition", 700, 100)
        cv2.imshow("Exit Recognition", frame)
        
        # cv2.imshow("Exit Recognition", frame)
        app.update_idletasks()
        app.update()

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    video_capture.release()
    cv2.destroyAllWindows()
    status_label.configure(text="🛑 سیستم تشخیص چهره برای خروج متوقف شد.", text_color="red")





###########################################################################################################


def show_main_menu():
    for widget in app.winfo_children():
        widget.destroy()

    app_l1 = CTkLabel(app, text="سیستم کنترل ورود و خروج", font=("Arial", 30, "bold"))
    app_l1.pack(pady=10)
    
    persian_days = [
        "دوشنبه",
        "سه‌شنبه",
        "چهارشنبه",
        "پنج‌شنبه",
        "جمعه",
        "شنبه",
        "یک‌شنبه"
    ]
    persian_months = [
        "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
        "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
    ]
    persian_digits = str.maketrans("0123456789", "۰۱۲۳۴۵۶۷۸۹")

    def to_persian_digits(s: str) -> str:
        return s.translate(persian_digits)

    def get_persian_day_name(dt: datetime.date) -> str:
        return persian_days[dt.weekday()]

    def get_jalali_date_text(dt: datetime.date, use_persian_digits=False) -> str:
        j = jdatetime.date.fromgregorian(date=dt)
        month_name = persian_months[j.month - 1]
        date_text = f"{j.day} {month_name} {j.year}"
        if use_persian_digits:
            return to_persian_digits(date_text)
        return date_text

    def get_time_text(now: datetime, use_persian_digits=False) -> str:
        t = now.strftime("%H:%M:%S")
        if use_persian_digits:
            return to_persian_digits(t)
        return t

    USE_PERSIAN_DIGITS = False  # True برای اعداد فارسی

    app_l2 = CTkLabel(app, text="", font=("Vazir", 20, "bold"))
    app_l2.pack(pady=50)

    # قبل از تعریف update_clock در show_main_menu بنویس:
    global clock_job

    # داخل update_clock:
    def update_clock():
        if not app_l2.winfo_exists():
            return
        global clock_job
        now_dt = datetime.now()
        today_date = now_dt.date()

        day_name = get_persian_day_name(today_date)
        jalali_text = get_jalali_date_text(today_date, use_persian_digits=USE_PERSIAN_DIGITS)
        time_text = get_time_text(now_dt, use_persian_digits=USE_PERSIAN_DIGITS)

        display = f"امروز: {day_name} | {jalali_text} | {time_text}"
        if app_l2.winfo_exists():  # ✅ فقط اگر هنوز وجود دارد
            app_l2.configure(text=display)
            clock_job = app.after(1000, update_clock)


    # 🔹 بدون این خط ساعت نمایش داده نمی‌شود:
    update_clock()

    

    option1 = StringVar(value="انتخاب کنید")
    app_menu = CTkOptionMenu(
        app,
        variable=option1,
        values=["ورود", "خروج", "پنل مدیر"],
        width=200,
        height=40,
        font=("Arial", 16)
    )
    app_menu.pack(pady=10)

    result_label = CTkLabel(app, text="", font=("Arial", 20))
    result_label.pack(pady=10)

    def confirm_mode():
        global stop_camera
        mode = option1.get()
        result_label.configure(text="")
        
        stop_camera = True
        cv2.destroyAllWindows()
        time.sleep(0.5)

        
        if mode == "ورود":
            result_label.configure(text="مد ورود انتخاب شد ✅")
            enter_cv2()
            # face_recognition_mode("entry")
        elif mode == "خروج":
            result_label.configure(text="مد خروج انتخاب شد 🕒")
            exit_cv2()
            # face_recognition_mode("exit")
        elif mode == "پنل مدیر":
            boss_panel()

    app_b1 = CTkButton(app, text="تأیید", command=confirm_mode, width=120, height=40, font=("Arial", 18))
    app_b1.pack(pady=10)
    
    


set_appearance_mode("dark")
set_default_color_theme("blue")

app = CTk()
app.title("سیستم کنترل ورود و خروج")
app.geometry("600x550+50+100")
init_db()
show_main_menu()
app.mainloop()


#############################################################################################################################



# import cv2
# import numpy as np
# import os
# import threading
# from datetime import datetime
# import pandas as pd
# from customtkinter import *
# from database import *

# # ---------------- تنظیمات ----------------
# CAMERA_URL = "http://192.168.1.2:8080/video"

# FRAME_REDUCTION = 0.25  # برای افزایش سرعت پردازش
# current_mode = None
# stop_camera = False

# # ---------------- GUI اصلی ----------------
# app = CTk()
# app.title("سیستم کنترل ورود و خروج")
# app.geometry("700x600")

# # ---------------- توابع کمکی ----------------
# def update_excel(full_name, mode, date_today=None, time_now=None):
#     excel_dir = os.path.join(os.path.dirname(__file__), "Excel Employees")
#     os.makedirs(excel_dir, exist_ok=True)
#     excel_path = os.path.join(excel_dir, f"{full_name}.xlsx")

#     if not date_today: date_today = datetime.now().strftime("%Y-%m-%d")
#     if not time_now: time_now = datetime.now().strftime("%H:%M:%S")

#     if os.path.exists(excel_path):
#         df = pd.read_excel(excel_path)
#     else:
#         df = pd.DataFrame(columns=["روز", "تاریخ", "ساعت ورود", "ساعت خروج"])

#     if mode == "entry":
#         df = pd.concat([df, pd.DataFrame([{"روز": datetime.now().strftime("%A"), "تاریخ": date_today, "ساعت ورود": time_now, "ساعت خروج": ""}])], ignore_index=True)
#     elif mode == "exit":
#         mask = df['تاریخ'] == date_today
#         df.loc[mask, 'ساعت خروج'] = time_now

#     df.to_excel(excel_path, index=False)

# # ---------------- تشخیص چهره ----------------
# import insightface
# model = insightface.app.FaceAnalysis(allowed_modules=['detection', 'recognition'])
# model.prepare(ctx_id=-1, det_size=(640, 640))  # CPU

# known_embeddings = []
# known_names = []

# # بارگذاری کارمندان و استخراج embedding
# for emp in get_all_employees():
#     emp_id, name, code, img_path = emp
#     if img_path and os.path.exists(img_path):
#         img = cv2.imread(img_path)
#         faces = model.get(img)
#         if faces:
#             known_embeddings.append(faces[0].embedding)
#             known_names.append(name)

# # ---------------- حلقه پردازش ----------------
# def face_recognition_loop(mode):
#     global stop_camera
#     stop_camera = False
#     video_capture = cv2.VideoCapture(CAMERA_URL)
#     detected_names = set()

#     while True:
#         if stop_camera: break
#         ret, frame = video_capture.read()
#         if not ret: continue

#         small_frame = cv2.resize(frame, (0,0), fx=FRAME_REDUCTION, fy=FRAME_REDUCTION)
#         faces = model.get(small_frame)

#         for face in faces:
#             embedding = face.embedding
#             distances = [np.linalg.norm(embedding - k) for k in known_embeddings]
#             if distances and min(distances) < 0.6:
#                 name = known_names[np.argmin(distances)]
#             else:
#                 name = "Unknown"

#             # ثبت ورود/خروج
#             if name != "Unknown" and name not in detected_names:
#                 if mode == "entry":
#                     mark_entry(name)
#                 else:
#                     mark_exit(name)
#                 update_excel(name, mode)
#                 detected_names.add(name)

#             # رسم روی فریم
#             x1, y1, x2, y2 = face.bbox.astype(int)
#             color = (0,255,0) if name != "Unknown" else (0,0,255)
#             cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
#             cv2.putText(frame, name, (x1, y2+20), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255,255,255),2)

#         cv2.imshow(f"{mode} Recognition", frame)
#         if cv2.waitKey(1) & 0xFF == ord('q'): break

#     video_capture.release()
#     cv2.destroyAllWindows()

# # ---------------- دکمه‌های GUI ----------------
# def start_entry():
#     threading.Thread(target=face_recognition_loop, args=("entry",)).start()

# def start_exit():
#     threading.Thread(target=face_recognition_loop, args=("exit",)).start()

# def stop_camera_btn():
#     global stop_camera
#     stop_camera = True

# CTkButton(app, text="ثبت ورود", command=start_entry, width=200).pack(pady=20)
# CTkButton(app, text="ثبت خروج", command=start_exit, width=200).pack(pady=20)
# CTkButton(app, text="توقف دوربین", command=stop_camera_btn, width=200).pack(pady=20)

# app.mainloop()
